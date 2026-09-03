# -*- coding: utf-8 -*-
"""
[NOMBRE POR DECIDIR] — Fotos a Excel, herramienta universal
Sube un Excel con una columna de referencia/SKU y tus fotos (ZIP o
sueltas). Devuelve el mismo Excel con la foto de cada referencia
incrustada en la primera columna. Funciona con cualquier catálogo,
no solo con nombres de archivo idénticos a la referencia.
"""

import io
import os
import re
import zipfile
import tempfile

import pandas as pd
import streamlit as st
from PIL import Image as PILImage
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

NOMBRE_APP = "Aparell"  # nombre definitivo

REF_CANDIDATAS = [
    "REF.", "REF", "REFERENCIA", "REFERÈNCIA", "REFERENCE",
    "COD", "CODIGO", "CÓDIGO", "CODE", "SKU", "ID", "PRODUCT ID",
    "ITEM", "ARTICULO", "ARTÍCULO", "PRODUCT",
]
COL_FOTOS = "FOTOS"
EXTS = (".jpg", ".jpeg", ".png", ".webp")
IMG_W, IMG_H, ROW_H, COL_A_W = 120, 120, 95, 20


def norm_header(s):
    s = str(s).strip().upper()
    return (s.replace("Á", "A").replace("É", "E").replace("Í", "I")
            .replace("Ó", "O").replace("Ú", "U").replace("Ü", "U"))


def detectar_col_ref(df):
    mapa = {norm_header(c): c for c in df.columns}
    for cand in REF_CANDIDATAS:
        c = norm_header(cand)
        if c in mapa:
            return mapa[c]
    return None


def normalizar_ref(ref):
    if ref is None:
        return None
    s = str(ref).strip().upper()
    if not s or s == "NAN":
        return None
    return s.replace("/", "_").replace("-", "_").replace(" ", "_")


def coincide_con_limite(ref, nombre_archivo):
    """¿Aparece 'ref' dentro de 'nombre_archivo' como unidad completa?
    Evita que REF1 case dentro de REF10, pero sí permite que
    'foto_REF001_frontal' encuentre a REF001."""
    idx = nombre_archivo.find(ref)
    while idx != -1:
        antes = nombre_archivo[idx - 1] if idx > 0 else ""
        despues = nombre_archivo[idx + len(ref)] if idx + len(ref) < len(nombre_archivo) else ""
        limite_antes = antes == "" or not antes.isalnum()
        limite_despues = despues == "" or not despues.isalnum()
        if limite_antes and limite_despues:
            return True
        idx = nombre_archivo.find(ref, idx + 1)
    return False


def preparar_imagen_para_excel(ruta):
    """Convierte cualquier imagen a JPEG en memoria antes de insertarla en
    el Excel. Evita que formatos como WEBP, o extensiones con mayúsculas,
    hagan fallar el guardado del Excel (openpyxl no los soporta bien)."""
    im = PILImage.open(ruta)
    if im.mode not in ("RGB", "L"):
        im = im.convert("RGB")
    buf = io.BytesIO()
    im.save(buf, format="JPEG", quality=90)
    buf.seek(0)
    return buf


def volcar_fotos(subidas, destino):
    for f in subidas:
        if f.name.lower().endswith(".zip"):
            try:
                with zipfile.ZipFile(f) as z:
                    z.extractall(destino)
            except zipfile.BadZipFile:
                st.error(f"El archivo {f.name} no es un ZIP válido.")
                st.stop()
        else:
            with open(os.path.join(destino, f.name), "wb") as d:
                d.write(f.getbuffer())


def indexar_fotos(carpeta):
    indice = {}
    for root, dirs, files in os.walk(carpeta):
        # Ignorar la carpeta de metadatos que macOS mete al comprimir ZIP
        dirs[:] = [d for d in dirs if d != "__MACOSX"]
        for f in files:
            # Ignorar archivos ocultos de metadatos de macOS (._foto.jpg),
            # que no son imágenes de verdad aunque tengan esa extensión.
            if f.startswith("._") or f.startswith("."):
                continue
            if f.lower().endswith(EXTS):
                nombre = os.path.splitext(f)[0].strip().upper()
                indice.setdefault(nombre, os.path.join(root, f))
    return indice


def buscar_foto(ref_norm, indice):
    """1) Coincidencia exacta (la mejor). 2) Si no hay, busca la ref
    como unidad dentro de cualquier nombre de archivo (coincidencia
    parcial, p.ej. 'foto_REF001_frontal.jpg')."""
    if ref_norm in indice:
        return indice[ref_norm]
    candidatas = sorted(
        nombre for nombre in indice if coincide_con_limite(ref_norm, nombre)
    )
    if candidatas:
        return indice[candidatas[0]]
    return None


def procesar(excel_bytes, carpeta_fotos, progreso):
    df = pd.read_excel(io.BytesIO(excel_bytes))

    col_ref = detectar_col_ref(df)
    if not col_ref:
        raise ValueError(
            "No encuentro la columna de referencia/SKU. "
            f"Columnas del Excel: {', '.join(str(c) for c in df.columns)}"
        )

    if COL_FOTOS in df.columns:
        df[COL_FOTOS] = ""
        df = df[[COL_FOTOS] + [c for c in df.columns if c != COL_FOTOS]]
    else:
        df.insert(0, COL_FOTOS, "")

    tmp_xlsx = os.path.join(carpeta_fotos, "_salida.xlsx")
    df.to_excel(tmp_xlsx, index=False)

    indice = indexar_fotos(carpeta_fotos)

    wb = load_workbook(tmp_xlsx)
    ws = wb.active
    ws.column_dimensions["A"].width = COL_A_W

    headers = {cell.value: i + 1 for i, cell in enumerate(ws[1])}
    col_ref_idx = headers[col_ref]

    fotos_ok, sin_foto = 0, []
    total = ws.max_row - 1
    for i, row in enumerate(range(2, ws.max_row + 1), start=1):
        ref_norm = normalizar_ref(ws.cell(row=row, column=col_ref_idx).value)
        ruta = buscar_foto(ref_norm, indice) if ref_norm else None
        insertada = False
        if ruta:
            try:
                buf_img = preparar_imagen_para_excel(ruta)
                img = XLImage(buf_img)
                img.width, img.height = IMG_W, IMG_H
                ws.row_dimensions[row].height = ROW_H
                ws.add_image(img, f"A{row}")
                fotos_ok += 1
                insertada = True
            except Exception:
                pass  # archivo no válido como imagen: se trata como "sin foto"
        if not insertada and ref_norm:
            sin_foto.append(ref_norm)
        if total:
            progreso(i / total)

    salida = io.BytesIO()
    wb.save(salida)
    salida.seek(0)
    return salida, col_ref, fotos_ok, total, sin_foto


# ---------- Interfaz ----------
st.set_page_config(page_title=f"{NOMBRE_APP} — fotos a Excel", page_icon="📊")

st.title(NOMBRE_APP)
st.caption("Sube un Excel con tu referencia/SKU y tus fotos. Te devuelve el mismo Excel con la foto de cada producto en la primera columna.")

st.info(
    "🔒 Tus archivos se procesan solo durante la generación y no se guardan "
    "en ningún servidor. Se pierden en cuanto cierras esta página.",
    icon="🔒",
)

excel_subido = st.file_uploader(
    "1 · Excel con la columna de referencia/SKU (REF, SKU, CODE, ID...)",
    type=["xlsx", "xlsm"],
)
fotos_subidas = st.file_uploader(
    "2 · Fotos: arrastra un ZIP con la carpeta, varios ZIP, fotos sueltas, o todo mezclado",
    type=["zip", "jpg", "jpeg", "png", "webp"],
    accept_multiple_files=True,
)

if excel_subido and fotos_subidas:
    if st.button("Generar Excel con fotos", type="primary"):
        with tempfile.TemporaryDirectory() as tmp:
            volcar_fotos(fotos_subidas, tmp)
            barra = st.progress(0.0, text="Incrustando fotos…")
            try:
                salida, col_ref, fotos_ok, total, sin_foto = procesar(
                    excel_subido.getvalue(), tmp,
                    lambda p: barra.progress(p, text=f"Incrustando fotos… {int(p*100)}%"),
                )
            except ValueError as e:
                st.error(str(e))
                st.stop()
            barra.empty()

        st.success(f"Listo: {fotos_ok} de {total} referencias con foto (columna detectada: {col_ref}).")
        if sin_foto:
            with st.expander(f"⚠ {len(sin_foto)} referencias sin foto"):
                st.text("\n".join(sin_foto))

        nombre = os.path.splitext(excel_subido.name)[0] + "_con_fotos.xlsx"
        st.download_button(
            "⬇ Descargar " + nombre, data=salida, file_name=nombre,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
else:
    st.info("Sube el Excel y al menos una foto o ZIP para empezar.")
